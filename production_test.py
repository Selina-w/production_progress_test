import streamlit as st
import pandas as pd
import matplotlib.pyplot as plt
import networkx as nx
import numpy as np
from datetime import datetime, timedelta
import matplotlib.font_manager as fm
import io
import tempfile
import os
import zipfile
import matplotlib as mpl
import json
import pathlib
import openpyxl
from openpyxl.styles import Font, Border, Alignment, PatternFill


# Create data directory if it doesn't exist
DATA_DIR = pathlib.Path("user_data")
DATA_DIR.mkdir(exist_ok=True)

def save_user_data(user_id, data):
    """Save user data to a JSON file"""
    user_file = DATA_DIR / f"{user_id}.json"
    with open(user_file, 'w', encoding='utf-8') as f:
        json.dump(data, f, default=str)

def load_user_data(user_id):
    """Load user data from JSON file"""
    user_file = DATA_DIR / f"{user_id}.json"
    if user_file.exists():
        with open(user_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
            # Convert string dates back to date objects
            for style in data.get("all_styles", []):
                style["sewing_start_date"] = datetime.strptime(style["sewing_start_date"], "%Y-%m-%d").date()
            return data
    return {"all_styles": []}
fm._load_fontmanager()
# Path relative to your script
font_path = os.path.join(os.path.dirname(__file__), "static", "simhei.ttf")
prop = fm.FontProperties(fname=font_path, size=22, weight='bold')
plt.rcParams['font.sans-serif'] = ['PingFang HK', 'Songti SC', 'SimHei', 'Arial Unicode MS']
plt.rcParams['font.family'] = 'sans-serif'
plt.rcParams['axes.unicode_minus'] = False  # Fix minus signs
plt.rcParams['figure.dpi'] = 300
plt.rcParams['savefig.dpi'] = 300
plt.rcParams['path.simplify'] = False  # Don't simplify paths for better quality
plt.rcParams['agg.path.chunksize'] = 10000  # Increase path chunk size
plt.rcParams['figure.facecolor'] = 'white'
plt.rcParams['figure.edgecolor'] = 'white'
plt.rcParams['lines.antialiased'] = True
plt.rcParams['patch.antialiased'] = True
plt.rcParams['text.antialiased'] = True
plt.rcParams['text.hinting'] = 'auto'  # Better text rendering
plt.rcParams['text.hinting_factor'] = 8  # Sharper text
plt.rcParams['text.usetex'] = False  # Disable LaTeX by default
plt.style.use('default')  # Reset to default style for clean rendering

fm._load_fontmanager()
# 检查字体是否可用
font_names = [f.name for f in fm.fontManager.ttflist]
chinese_fonts = [f for f in font_names if any(name in f for name in ['PingFang', 'Microsoft', 'SimHei', 'Arial Unicode'])]
if chinese_fonts:
    plt.rcParams['font.sans-serif'] = chinese_fonts[0]
    print(chinese_fonts[0])

# 部门工序定义
def get_department_steps(process_type=None):
    """Get department steps based on process type"""
    all_departments = {
        "产前确认": ["代用面料裁剪", "满花样品", "局花样品", "绣花样品", "版型", "代用样品发送", "版型确认", 
                 "印绣样品确认", "辅料样发送", "辅料确认", "色样发送", "色样确认"],
        "面料": ["仕样书", "工艺分析", "排版", "用料", "棉纱", "毛坯", "光坯", "物理检测验布"],
        "满花": ["满花工艺", "满花", "满花后整", "物理检测"],
        "裁剪": ["工艺样版", "裁剪"],
        "局花": ["局花工艺", "局花", "物理检测"],
        "绣花": ["绣花工艺", "绣花", "物理检测"],
        "配片": ["配片"],
        "滚领": ["滚领布"],
        "辅料": ["辅料限额", "辅料", "物理检测"],
        "缝纫": ["缝纫工艺", "缝纫开始", "缝纫结束"],
        "后整": ["后整工艺", "检验", "包装", "检针装箱"],
        "工艺": ["船样检测摄影", "外观"]
    }
    
    # If no process type is specified, return all departments
    if process_type is None:
        return all_departments

    # Define exclusions for different process types
    exclusion_map = {
        "满花局花": ["绣花"],
        "满花绣花": ["局花"],
        "局花绣花": ["满花"],
        "满花": ["局花", "绣花"],
        "局花": ["满花", "绣花"],
        "绣花": ["满花", "局花"]
    }

    # If process_type is invalid, return all departments
    if process_type not in exclusion_map and process_type != "满花局花绣花":
        raise ValueError(f"Invalid process_type: {process_type}")

    # Exclude specified departments
    if process_type in exclusion_map:
        filtered_departments = {k: v for k, v in all_departments.items() if k not in exclusion_map[process_type]}
        # Remove corresponding steps from "产前确认"
        if "产前确认" in filtered_departments:
            remove_steps = {"满花": "满花样品", "局花": "局花样品", "绣花": "绣花样品"}
            filtered_departments["产前确认"] = [step for step in filtered_departments["产前确认"]
                                                if step not in [remove_steps.get(p) for p in exclusion_map[process_type] if p in remove_steps]]
        return filtered_departments

    # If "满花局花绣花", return all departments
    return all_departments

def calculate_schedule(sewing_start_date, process_type, confirmation_period, order_quantity, daily_production, start_time_period="上午"):
    """ 计算整个生产流程的时间安排 """
    schedule = {}
    
    # 将所有工序的时间初始化为字典
    for dept, steps in get_department_steps(process_type).items():
        schedule[dept] = {}
    
    Y = sewing_start_date  # 订单缝纫开始日期
    if confirmation_period == 7:
        if process_type == "满花局花绣花":
            X = Y - timedelta(days=54)
        elif process_type == "满花局花":
            X = Y - timedelta(days=47)
        elif process_type == "满花绣花":
            X = Y - timedelta(days=49)
        elif process_type == "局花绣花":
            X = Y - timedelta(days=48)
        elif process_type == "满花":
            X = Y - timedelta(days=42)
        elif process_type == "局花":
            X = Y - timedelta(days=41)
        else:
            X = Y - timedelta(days=43)

    elif confirmation_period == 14:
        if process_type == "满花局花绣花":
            X = Y - timedelta(days=61)
        elif process_type == "满花局花":
            X = Y - timedelta(days=54)
        elif process_type == "满花绣花":
            X = Y - timedelta(days=56)
        elif process_type == "局花绣花":
            X = Y - timedelta(days=55)
        elif process_type == "满花":
            X = Y - timedelta(days=49)
        elif process_type == "局花":
            X = Y - timedelta(days=48)
        else:
            X = Y - timedelta(days=50)

    elif confirmation_period == 30:
        if process_type == "满花局花绣花":
            X = Y - timedelta(days=77)
        elif process_type == "满花局花":
            X = Y - timedelta(days=70)
        elif process_type == "满花绣花":
            X = Y - timedelta(days=72)
        elif process_type == "局花绣花":
            X = Y - timedelta(days=68)
        elif process_type == "满花":
            X = Y - timedelta(days=65)
        elif process_type == "局花":
            X = Y - timedelta(days=61)
        else:
            X = Y - timedelta(days=63)
    # 1. 计算产前确认阶段
    schedule["产前确认"]["代用面料裁剪"] = {"时间点": X + timedelta(days=20)}
    if "满花" in process_type:
        schedule["产前确认"]["满花样品"] = {"时间点": X + timedelta(days=23)}
    if process_type == "满花局花绣花" or process_type == "满花局花":
        schedule["产前确认"]["局花样品"] = {"时间点": X + timedelta(days=24)}
    else:
        if "局花" in process_type:
            schedule["产前确认"]["局花样品"] = {"时间点": X + timedelta(days=23)}
    if process_type == "满花局花绣花":
        schedule["产前确认"]["绣花样品"] = {"时间点": X + timedelta(days=25)}
        schedule["产前确认"]["版型"] = {"时间点": X + timedelta(days=27)}
        schedule["产前确认"]["代用样品发送"] = {"时间点": X + timedelta(days=28)}
    elif process_type == "满花局花":
        schedule["产前确认"]["版型"] = {"时间点": X + timedelta(days=26)}
        schedule["产前确认"]["代用样品发送"] = {"时间点": X + timedelta(days=27)}
    elif process_type == "满花绣花" or process_type == "局花绣花":
        schedule["产前确认"]["绣花样品"] = {"时间点": X + timedelta(days=24)}
        schedule["产前确认"]["版型"] = {"时间点": X + timedelta(days=26)}
        schedule["产前确认"]["代用样品发送"] = {"时间点": X + timedelta(days=27)}
    elif process_type == "满花" or process_type == "局花":
        schedule["产前确认"]["版型"] = {"时间点": X + timedelta(days=25)}
        schedule["产前确认"]["代用样品发送"] = {"时间点": X + timedelta(days=26)}
    elif process_type == "绣花":
        schedule["产前确认"]["绣花样品"] = {"时间点": X + timedelta(days=23)}
        schedule["产前确认"]["版型"] = {"时间点": X + timedelta(days=25)}
        schedule["产前确认"]["代用样品发送"] = {"时间点": X + timedelta(days=26)}
    if confirmation_period == 30:
        schedule["产前确认"]["版型确认"] = {"时间点": schedule["产前确认"]["代用样品发送"]["时间点"] + timedelta(days=20)}
        schedule["产前确认"]["印绣样品确认"] = {"时间点": schedule["产前确认"]["代用样品发送"]["时间点"] + timedelta(days=confirmation_period)}
    else:
        schedule["产前确认"]["版型确认"] = {"时间点": schedule["产前确认"]["代用样品发送"]["时间点"] + timedelta(days=confirmation_period)}
        schedule["产前确认"]["印绣样品确认"] = {"时间点": schedule["产前确认"]["版型确认"]["时间点"]}
    schedule["产前确认"]["辅料样发送"] = {"时间点": X + timedelta(days=27)}
    if confirmation_period == 30:
        schedule["产前确认"]["辅料确认"] = {"时间点": schedule["产前确认"]["辅料样发送"]["时间点"] + timedelta(days=20)}
        schedule["产前确认"]["色样发送"] = {"时间点": X + timedelta(days=15)}
        schedule["产前确认"]["色样确认"] = {"时间点": schedule["产前确认"]["色样发送"]["时间点"] + timedelta(days=20)}
    else:
        schedule["产前确认"]["辅料确认"] = {"时间点": schedule["产前确认"]["辅料样发送"]["时间点"] + timedelta(days=confirmation_period)}
        schedule["产前确认"]["色样发送"] = {"时间点": X + timedelta(days=15)}
        schedule["产前确认"]["色样确认"] = {"时间点": schedule["产前确认"]["色样发送"]["时间点"] + timedelta(days=confirmation_period)}
    
    # 2. 计算面料阶段
    schedule["面料"]["仕样书"] = {"时间点": X + timedelta(days=10)}
    schedule["面料"]["工艺分析"] = {"时间点": X + timedelta(days=11)}
    schedule["面料"]["排版"] = {"时间点": X + timedelta(days=12)}
    schedule["面料"]["用料"] = {"时间点": X + timedelta(days=12)}
    if confirmation_period == 7:
        schedule["面料"]["棉纱"] = {"时间点": X + timedelta(days=15)}
        schedule["面料"]["毛坯"] = {"时间点": X + timedelta(days=19)}
        schedule["面料"]["光坯"] = {"时间点": X + timedelta(days=27)}
    elif confirmation_period == 14:
        schedule["面料"]["棉纱"] = {"时间点": X + timedelta(days=16)}
        schedule["面料"]["毛坯"] = {"时间点": X + timedelta(days=21)}
        schedule["面料"]["光坯"] = {"时间点": X + timedelta(days=34)}
    elif confirmation_period == 30:
        schedule["面料"]["棉纱"] = {"时间点": X + timedelta(days=16)}
        schedule["面料"]["毛坯"] = {"时间点": X + timedelta(days=22)}
        schedule["面料"]["光坯"] = {"时间点": X + timedelta(days=40)}
    schedule["面料"]["物理检测验布"] = {"时间点": schedule["面料"]["光坯"]["时间点"] + timedelta(days=1)}

    # 3. 计算满花流程
    if confirmation_period == 7 or confirmation_period == 14:
        if process_type == "满花局花绣花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=7)}
        if process_type == "满花局花" or process_type == "满花绣花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=6)}
        if process_type == "满花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=5)}
    elif confirmation_period == 30:
        if process_type == "满花局花绣花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=17)}
        if process_type == "满花局花" or process_type == "满花绣花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=16)}
        if process_type == "满花":
            schedule["满花"]["满花工艺"] = {"时间点": schedule["面料"]["物理检测验布"]["时间点"]+ timedelta(days=15)}
    
    if "满花" in schedule:
        schedule["满花"]["满花"] = {"时间点": schedule["满花"]["满花工艺"]["时间点"] + timedelta(days=3)}
        schedule["满花"]["满花后整"] = {"时间点": schedule["满花"]["满花"]["时间点"] + timedelta(days=1)}
        schedule["满花"]["物理检测"] = {"时间点": schedule["满花"]["满花后整"]["时间点"] + timedelta(days=1)}

    # 5. 计算裁剪流程
    schedule["裁剪"]["工艺样版"] = {"时间点": schedule["产前确认"]["版型确认"]["时间点"]}
    
    if process_type == "局花" or process_type == "绣花" or process_type == "局花绣花":
        schedule["裁剪"]["裁剪"] = {"时间点": schedule["裁剪"]["工艺样版"]["时间点"] + timedelta(days=3)}
    else:
        schedule["裁剪"]["裁剪"] = {"时间点": schedule["满花"]["物理检测"]["时间点"] + timedelta(days=3)}


    # 4. 计算局花流程
    if "局花" in schedule:
        schedule["局花"]["局花工艺"] = {"时间点": schedule["产前确认"]["印绣样品确认"]["时间点"]}
        if process_type == "满花局花绣花" or process_type == "满花局花":
            schedule["局花"]["局花"] = {"时间点": schedule["局花"]["局花工艺"]["时间点"] + timedelta(days=11)}
        else:
            if confirmation_period == 30:
                schedule["局花"]["局花"] = {"时间点": schedule["局花"]["局花工艺"]["时间点"] + timedelta(days=3)}
            else:
                schedule["局花"]["局花"] = {"时间点": schedule["局花"]["局花工艺"]["时间点"] + timedelta(days=6)}
        schedule["局花"]["物理检测"] = {"时间点": schedule["局花"]["局花"]["时间点"] + timedelta(days=1)}

    # 5. 计算绣花流程
    if "绣花" in schedule:
        if confirmation_period == 30:
            schedule["绣花"]["绣花工艺"] = {"时间点": schedule["裁剪"]["工艺样版"]["时间点"] + timedelta(days=10)}
            if process_type == "满花局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":schedule["局花"]["物理检测"]["时间点"] + timedelta(days=5)}
            elif process_type == "满花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=70)}
            elif process_type == "局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=66)}
            else:
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=61)}

        elif confirmation_period == 7:
            schedule["绣花"]["绣花工艺"] = {"时间点": schedule["裁剪"]["工艺样版"]["时间点"]}
            if process_type == "满花局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":schedule["局花"]["物理检测"]["时间点"] + timedelta(days=5)}
            elif process_type == "满花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=47)}
            elif process_type == "局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=46)}
            else:
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=41)}
        elif confirmation_period == 14:
            schedule["绣花"]["绣花工艺"] = {"时间点": schedule["裁剪"]["工艺样版"]["时间点"]}
            if process_type == "满花局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=59)}
            elif process_type == "满花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=54)}
            elif process_type == "局花绣花":
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=53)}
            else:
                schedule["绣花"]["绣花"] = {"时间点":X + timedelta(days=48)}
        schedule["绣花"]["物理检测"] = {"时间点": schedule["绣花"]["绣花"]["时间点"] + timedelta(days=1)}

    # 6. 计算配片
    if "绣花" in schedule:
        schedule["配片"]["配片"] = {
                "时间点": schedule["绣花"]["物理检测"]["时间点"]
            }
    else:
        if "局花" in schedule:
            schedule["配片"]["配片"] = {"时间点": schedule["局花"]["物理检测"]["时间点"]}
        else:
            schedule["配片"]["配片"] = {"时间点": schedule["裁剪"]["裁剪"]["时间点"]}

    # 7. 计算滚领
    schedule["滚领"]["滚领布"] = {"时间点": schedule["配片"]["配片"]["时间点"]}

    # 8. 计算辅料流程（并行）--从这开始
    schedule["辅料"]["辅料限额"] = {"时间点":  X + timedelta(days=17)}
    if confirmation_period == 7:
        if process_type == "满花局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=49)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=53)}
        elif process_type == "满花局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=45)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=45)}
        elif process_type == "满花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=47)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=47)}
        elif process_type == "局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=46)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=46)}
        elif process_type == "满花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=40)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=40)}
        elif process_type == "局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=39)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=39)}
        else:
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=41)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=41)}

    elif confirmation_period == 14:
        if process_type == "满花局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=55)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=59)}
        elif process_type == "满花局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=52)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=52)}
        elif process_type == "满花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=54)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=54)}
        elif process_type == "局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=53)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=53)}
        elif process_type == "满花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=47)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=47)}
        elif process_type == "局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=46)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=46)}
        else:
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=48)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=48)}

    elif confirmation_period == 30:
        if process_type == "满花局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=62)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=75)}
        elif process_type == "满花局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=62)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=68)}
        elif process_type == "满花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=62)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=70)}
        elif process_type == "局花绣花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=62)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=66)}
        elif process_type == "满花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=62)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=63)}
        elif process_type == "局花":
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=59)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=59)}
        else:
            schedule["辅料"]["辅料"] = {"时间点":X + timedelta(days=61)}
            schedule["缝纫"]["缝纫工艺"] = {"时间点": X + timedelta(days=61)}

    schedule["辅料"]["物理检测"] = {"时间点": schedule["辅料"]["辅料"]["时间点"] + timedelta(days=1)}

    # 9. 计算缝纫工艺
    schedule["缝纫"]["缝纫开始"] = {"时间点": sewing_start_date, "备注": start_time_period} #{"时间点": schedule["缝纫"]["缝纫工艺"]["时间点"] + timedelta(days=2)}
    
    # 计算缝纫结束时间，根据小数部分决定是当天上午结束还是下午结束或第二天
    sewing_days_float = order_quantity * 1.05 / daily_production
    sewing_days_int = int(sewing_days_float)
    sewing_days_decimal = sewing_days_float - sewing_days_int
    
    if sewing_days_decimal <= 0.5:
        # 如果小数部分小于等于0.5，则当天中午结束
        sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int)
        schedule["缝纫"]["缝纫结束"] = {
            "时间点": sewing_end_date,
            "备注": "中午结束" if sewing_days_decimal > 0 else "全天"
        }
    else:
        # 如果小数部分大于0.5，则需要多一天
        sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int + 1)
        schedule["缝纫"]["缝纫结束"] = {
            "时间点": sewing_end_date,
            "备注": "全天"
        }

    # 考虑开始时间是上午还是下午
    if start_time_period == "上午":
        if sewing_days_decimal <= 0.5:
            # 如果小数部分小于等于0.5，则当天下午结束
            sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int)
            schedule["缝纫"]["缝纫结束"] = {
                "时间点": sewing_end_date,
                "备注": "下午" if sewing_days_decimal > 0 else "上午"
            }
        else:
            # 如果小数部分大于0.5，则第二天上午结束
            sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int + 1)
            schedule["缝纫"]["缝纫结束"] = {
                "时间点": sewing_end_date,
                "备注": "上午"
            }
    else:  # 下午开始
        if sewing_days_decimal <= 0:
            # 如果刚好整数天，则最后一天下午结束
            sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int)
            schedule["缝纫"]["缝纫结束"] = {
                "时间点": sewing_end_date,
                "备注": "下午"
            }
        elif sewing_days_decimal <= 0.5:
            # 如果小数部分小于等于0.5，则第二天上午结束
            sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int + 1)
            schedule["缝纫"]["缝纫结束"] = {
                "时间点": sewing_end_date,
                "备注": "上午"
            }
        else:
            # 如果小数部分大于0.5，则第二天下午结束
            sewing_end_date = schedule["缝纫"]["缝纫开始"]["时间点"] + timedelta(days=sewing_days_int + 1)
            schedule["缝纫"]["缝纫结束"] = {
                "时间点": sewing_end_date,
                "备注": "下午"
            }


    # 10. 计算后整工艺
    schedule["后整"]["后整工艺"] = {"时间点": schedule["缝纫"]["缝纫工艺"]["时间点"]}
    schedule["后整"]["检验"] = {"时间点": schedule["缝纫"]["缝纫结束"]["时间点"]+ timedelta(days=1)}
    schedule["后整"]["包装"] = {"时间点": schedule["缝纫"]["缝纫结束"]["时间点"]+ timedelta(days=2)}
    schedule["后整"]["检针装箱"] = {"时间点": schedule["缝纫"]["缝纫结束"]["时间点"]+ timedelta(days=3)}

    # 11. 计算工艺
    schedule["工艺"]["船样检测摄影"] = {"时间点": schedule["后整"]["后整工艺"]["时间点"]+ timedelta(days=4)}
    schedule["工艺"]["检验"] = {"时间点": schedule["工艺"]["船样检测摄影"]["时间点"]+ timedelta(days=3)}
    

    return schedule

# 重新安排生产组中款式的缝纫开始时间
def rearrange_styles_by_production_group(styles):
    """
    重新安排同一生产组内款式的缝纫开始时间
    确保同一生产顺序的款式共享相同的开始时间
    确保下一个生产顺序的款式开始时间等于前一个生产顺序中最后一个款式的结束时间
    """
    # 将款式按生产组分组
    grouped_styles = {}
    for style in styles:
        group = style.get("production_group", "")
        if not group:  # 如果没有生产组，跳过
            continue
        
        if group not in grouped_styles:
            grouped_styles[group] = []
        grouped_styles[group].append(style)
    
    # 对每个生产组内的款式进行处理
    rearranged_styles = []
    for group, group_styles in grouped_styles.items():
        # 按照生产顺序进一步分组
        order_grouped_styles = {}
        for style in group_styles:
            order = style.get("production_order", 9999)
            if order not in order_grouped_styles:
                order_grouped_styles[order] = []
            order_grouped_styles[order].append(style)
        
        # 按生产顺序排序
        sorted_orders = sorted(order_grouped_styles.keys())
        
        # 处理第一个生产顺序组 - 保持原始开始日期
        first_order = sorted_orders[0]
        first_order_styles = order_grouped_styles[first_order]
        
        # 获取第一个款式的缝纫开始日期和时段作为这个组的共同开始时间
        if first_order_styles:
            # 可以按照日期排序，以确保使用最早的日期
            first_order_styles.sort(key=lambda x: x["sewing_start_date"])
            first_style = first_order_styles[0]
            start_date = first_style["sewing_start_date"]
            start_time_period = first_style.get("start_time_period", "上午")
            
            # 将相同开始时间应用于该组中的所有款式
            for style in first_order_styles:
                style["sewing_start_date"] = start_date
                style["start_time_period"] = start_time_period
                rearranged_styles.append(style)
            
            # 计算该组最后结束的时间 - 需要检查每个款式的结束时间
            latest_end_time = None
            latest_end_remark = "下午结束"
            
            for style in first_order_styles:
                sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
                schedule = calculate_schedule(
                    sewing_start_time, 
                    style["process_type"], 
                    style["cycle"], 
                    style["order_quantity"], 
                    style["daily_production"],
                    style["start_time_period"]
                )
                
                end_time = schedule["缝纫"]["缝纫结束"]["时间点"]
                end_remark = schedule["缝纫"]["缝纫结束"].get("备注", "下午结束")
                
                if latest_end_time is None or end_time > latest_end_time:
                    latest_end_time = end_time
                    latest_end_remark = end_remark
            
            # 依次处理后续生产顺序组
            for i in range(1, len(sorted_orders)):
                current_order = sorted_orders[i]
                current_order_styles = order_grouped_styles[current_order]
                
                # 前一个组的结束时间作为当前组的开始时间
                start_date = latest_end_time.date()
                
                # 根据前一个结束时段确定当前组的开始时段
                if "上午" in latest_end_remark:
                    start_time_period = "上午"
                else:
                    start_time_period = "下午"
                
                # 将相同开始时间应用于该组中的所有款式
                for style in current_order_styles:
                    style["sewing_start_date"] = start_date
                    style["start_time_period"] = start_time_period
                    rearranged_styles.append(style)
                
                # 更新最晚结束时间以供下一个组使用
                latest_end_time = None
                latest_end_remark = "下午结束"
                
                for style in current_order_styles:
                    sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
                    schedule = calculate_schedule(
                        sewing_start_time, 
                        style["process_type"], 
                        style["cycle"], 
                        style["order_quantity"], 
                        style["daily_production"],
                        style["start_time_period"]
                    )
                    
                    end_time = schedule["缝纫"]["缝纫结束"]["时间点"]
                    end_remark = schedule["缝纫"]["缝纫结束"].get("备注", "下午结束")
                    
                    if latest_end_time is None or end_time > latest_end_time:
                        latest_end_time = end_time
                        latest_end_remark = end_remark
    
    # 添加没有生产组的款式
    for style in styles:
        if not style.get("production_group", ""):
            rearranged_styles.append(style)
    
    return rearranged_styles

def generate_excel_report(styles):
    """生成包含所有款式信息的Excel报表，以日期为列，款号为行"""
    # 创建一个临时目录
    temp_dir = tempfile.mkdtemp()
    
    # 收集所有日期和步骤信息
    all_dates = set()
    style_steps = {}
    
    # 处理每个款式
    for style in styles:
        style_number = style["style_number"]
        style_steps[style_number] = {}
        
        # 如果款式已经有计算好的schedule，使用它
        if "schedule" in style:
            schedule = style["schedule"]
        else:
            # 否则重新计算schedule
            sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time()) if not isinstance(style["sewing_start_date"], datetime) else style["sewing_start_date"]
            schedule = calculate_schedule(
                sewing_start_time, 
                style["process_type"], 
                style["cycle"], 
                style["order_quantity"], 
                style["daily_production"],
                style.get("start_time_period", "上午")
            )
        
        # 收集每个步骤的日期和备注
        for dept, steps in schedule.items():
            for step, info in steps.items():
                time_point = info["时间点"]
                if hasattr(time_point, "date"):
                    date = time_point.date()
                else:
                    date = time_point
                
                all_dates.add(date)
                
                # 对于缝纫步骤，添加生产组信息
                if dept == "缝纫":
                    step_info = f"{dept}-{step}"
                    if style.get("production_group"):
                        step_info += f" ({style['production_group']})"
                else:
                    step_info = f"{dept}-{step}"
                
                # 如果有备注，添加到步骤信息中
                if "备注" in info:
                    step_info += f" [{info['备注']}]"
                
                # 如果同一天有多个步骤，用换行符分隔
                if date in style_steps[style_number]:
                    style_steps[style_number][date] += f"\n{step_info}"
                else:
                    style_steps[style_number][date] = step_info
    
    # 生成连续的日期序列
    min_date = min(all_dates)
    max_date = max(all_dates)
    all_dates = []
    current_date = min_date
    while current_date <= max_date:
        all_dates.append(current_date)
        current_date += timedelta(days=1)
    
    # 创建DataFrame
    data = []
    for style_number in sorted(style_steps.keys()):
        row = {"款号": style_number}
        for date in all_dates:
            row[date] = style_steps[style_number].get(date, "")
        data.append(row)
    
    df = pd.DataFrame(data)
    
    # 保存为Excel文件
    excel_path = os.path.join(temp_dir, "生产计划报表.xlsx")
    
    # 创建Excel写入器
    writer = pd.ExcelWriter(excel_path, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='生产计划')
    
    # 获取工作簿和工作表
    workbook = writer.book
    worksheet = writer.sheets['生产计划']
    
    # 定义边框样式
    thin_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thin')
    )
    
    # 添加标题行
    title_cell = worksheet['A1']
    title_cell.value = "生产计划跟踪记录"
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df.columns))
    title_cell.font = openpyxl.styles.Font(bold=True, size=14)
    title_cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='center')
    
    # # 将数据行向下移动一行
    # for row in range(len(df) + 1, 0, -1):
    #     for col in range(1, len(df.columns) + 1):
    #         cell = worksheet.cell(row=row, column=col)
    #         target_cell = worksheet.cell(row=row + 1, column=col)
    #         target_cell.value = cell.value
    #         if cell.has_style:
    #             if isinstance(cell.font, Font):
    #                 target_cell.font = cell.font
    #             if isinstance(cell.border, Border):
    #                 target_cell.border = cell.border
    #             if isinstance(cell.alignment, Alignment):
    #                 target_cell.alignment = cell.alignment
    #             if isinstance(cell.fill, PatternFill):
    #                 target_cell.fill = cell.fill
    #             target_cell.number_format = cell.number_format
    
    # 设置列宽和自动换行
    for i, col in enumerate(df.columns):
        # 获取Excel列引用
        col_letter = openpyxl.utils.get_column_letter(i + 1)
        
        # 设置列宽
        if col == "款号":
            column_width = max(len(str(col)), df[col].astype(str).map(len).max())
        else:
            column_width = 15  # 固定日期列的宽度
        worksheet.column_dimensions[col_letter].width = min(column_width + 2, 30)
        
        # 设置自动换行和边框
        for row in range(1, len(df) + 3):  # +3 because Excel is 1-based and we added a title row
            cell = worksheet[f"{col_letter}{row}"]
            cell.border = thin_border
            
            # 设置对齐方式
            if row == 2:  # 表头行 (now row 2 because of title)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
            elif row == 1:  # 标题行
                continue  # Skip title row as it's already formatted
            elif col == "款号":  # 款号列
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
            else:  # 日期列
                cell.alignment = openpyxl.styles.Alignment(horizontal='left', vertical='top', wrap_text=True)
    
    # 冻结首行和款号列
    worksheet.freeze_panes = 'B3'  # Changed from B2 to B3 to account for title row
    
    # 保存并关闭Excel文件
    writer.close()
    
    return excel_path
        
# 画时间线
def plot_timeline(schedule, process_type, confirmation_period):
    # 根据工序类型定义部门顺序和颜色
    if process_type == "满花局花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "局花", "裁剪", "满花", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "满花": "#C1E1FF", 
            "裁剪": "#D1FFC1", 
            "局花": "#FFC1E1", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    elif process_type == "满花绣花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "绣花", "裁剪", "满花", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "满花": "#C1E1FF", 
            "裁剪": "#D1FFC1", 
            "绣花": "#E1C1FF", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    elif process_type == "局花绣花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "绣花", "局花", "裁剪", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "裁剪": "#D1FFC1", 
            "绣花": "#E1C1FF", 
            "局花": "#FFC1E1", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    elif process_type == "满花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "裁剪", "满花", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "满花": "#C1E1FF", 
            "裁剪": "#D1FFC1", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    elif process_type == "局花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "局花", "裁剪", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "裁剪": "#D1FFC1", 
            "局花": "#FFC1E1", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    elif process_type == "绣花":
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "绣花", "裁剪", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "裁剪": "#D1FFC1", 
            "绣花": "#E1C1FF", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    else:  # "满花局花绣花"
        department_order = ["工艺", "后整", "缝纫", "辅料", "滚领", "配片", "绣花", "局花", "裁剪", "满花", "面料", "产前确认"]
        department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "满花": "#C1E1FF", 
            "裁剪": "#D1FFC1", 
            "局花": "#FFC1E1", 
            "绣花": "#E1C1FF", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    
    # 计算时间范围（不包括缝纫部分）
    min_date = min(times["时间点"] for dept in schedule for times in schedule[dept].values())
    max_date = max(times["时间点"] for dept in department_order for times in schedule[dept].values())
    date_range = (max_date - min_date).days
    
    # Calculate figure size and DPI based on date range
    base_width = int(date_range/41*40)  # Base width calculation
    if base_width > 24:  # If the figure is larger than default
        dpi_scale = base_width / 24  # Calculate how much larger it is
        plt.rcParams['figure.dpi'] = int(300 * dpi_scale)  # Scale DPI proportionally
        plt.rcParams['savefig.dpi'] = int(300 * dpi_scale)
    
    # Create figure with dynamic sizing and high-quality settings
    fig, ax = plt.subplots(figsize=(base_width, 25))
    fig.patch.set_facecolor('white')
    ax.set_facecolor('white')
    
    # Enable high-quality rendering
    ax.set_rasterization_zorder(1)
    plt.rcParams['svg.fonttype'] = 'none'  # Prevent text conversion to paths
    
    # 设置y轴位置，增加行间距
    y_positions = {dept: idx * 1.5 for idx, dept in enumerate(department_order, start=1)}
    
    # 设置主时间线的起始和结束位置（留出两端的空间）
    timeline_start = 0.02  # 2% margin from left
    fixed_end = 0.89      # 固定最晚时间点的位置
    timeline_width = fixed_end - timeline_start
    
    # 绘制主要部门
    for dept in department_order:
        y = y_positions[dept]
        steps = schedule[dept]
        step_names = list(steps.keys())
        step_dates = [steps[step]["时间点"] for step in step_names]
        
        # 创建部门背景 - 从0到0.91（保持原来的位置）
        ax.fill_betweenx([y - 0.4, y + 0.4], 0, 0.91, 
                        color=department_colors.get(dept, "#DDDDDD"), alpha=0.5)
        
        # 处理同一时间点的步骤
        time_groups = {}
        for i, date in enumerate(step_dates):
            # 将时间点映射到新的范围内，保持时间比例
            days_from_start = (date - min_date).days
            scaled_position = timeline_start + (days_from_start / date_range) * timeline_width
            
            if scaled_position not in time_groups:
                time_groups[scaled_position] = []
            time_groups[scaled_position].append((step_names[i], date))
        
        # 绘制步骤
        sorted_positions = sorted(time_groups.keys())
        for idx, x_pos in enumerate(sorted_positions):
            steps_at_time = time_groups[x_pos]
            # 对于同一时间点的所有步骤，只画一个点
            ax.scatter(x_pos, y, color='black', zorder=3)
            
            # 获取这个时间点的所有步骤的日期
            current_date = steps_at_time[0][1]
            
            # 检查前后是否有相邻的日期点
            prev_date = None if idx == 0 else time_groups[sorted_positions[idx-1]][0][1]
            next_date = None if idx == len(sorted_positions)-1 else time_groups[sorted_positions[idx+1]][0][1]
            
            # 垂直排列同一时间点的步骤
            for i, (step, date) in enumerate(steps_at_time):
                # 计算文本框的水平位置
                text_x = x_pos
                
                # 如果与前一个点相差一天，向左偏移文本
                if prev_date and abs((date - prev_date).days) == 1:
                    if confirmation_period == 7:
                        text_x = x_pos + 0.0135
                    else:
                        text_x = x_pos + 0.0105
                # 如果与后一个点相差一天，向右偏移文本
                elif next_date and abs((date - next_date).days) == 1:
                    if confirmation_period == 7:
                        text_x = x_pos - 0.0135
                    else:
                        text_x = x_pos - 0.0105
                
                # 计算垂直偏移，对于同一天的步骤使用垂直堆叠
                if i == 0:
                    y_offset = -0.3  # 第一个步骤的偏移保持不变
                else:
                    # 如果是同一天的步骤，垂直堆叠
                    if date == current_date:
                        y_offset = -0.3 - i * 0.52
                    else:
                        # 如果是不同天的步骤，保持相同的垂直位置
                        y_offset = -0.3
                
                # 绘制文本框
                text_box = dict(boxstyle='round,pad=0.4', 
                              facecolor='white', 
                              alpha=1.0,  # Full opacity for sharper text
                              edgecolor='black', 
                              linewidth=1,
                              snap=True)  # Snap to pixel grid
                # 在点下方显示步骤名称和日期，调整字体大小
                step_text = f"{step}\n{date.strftime('%Y/%m/%d')}"
                
                # 如果有备注信息，添加到文本中（针对缝纫结束时间）
                if dept == "缝纫" and step in ["缝纫结束", "缝纫开始"] and "备注" in schedule[dept][step]:
                    step_text = f"{step}\n{date.strftime('%Y/%m/%d')}\n{schedule[dept][step]['备注']}"
                
                # 特殊处理印布的印布后整，将其放在时间线上方
                if ((dept == "产前确认" and (step == "色样确认" or step == "绣花样品"))
                    or (dept == "面料" and step == "工艺分析")
                    or (dept == "满花" and step == "满花后整")
                    or (dept == "后整" and step == "包装")):
                    y_offset = 0.3  # Place above the timeline
                
                 # 1. 对于包含"满花"的流程（除了"满花绣花"）：将"满花样品"放到时间线上方
                if dept == "产前确认" and step == "满花样品" and "满花" in process_type and process_type != "满花绣花":
                    y_offset = 0.3  # 放在时间线上方
                
                # 2. 在"满花局花绣花"的情况下：将"菊花样品"放到时间线上方，并与时间线保持一个文本框的距离
                if dept == "产前确认" and step == "局花样品" and process_type == "满花局花绣花":
                    y_offset = 0.8  # 放在时间线上方，有更大的距离
                
                # 3. 除了"满花局花绣花"或"满花"的情况下：将"版型"步骤放到时间线下方，与时间线有一个文本框的距离
                if dept == "产前确认" and step == "版型" and process_type != "满花局花绣花" and process_type != "满花" and process_type != "绣花" and process_type != "局花绣花":
                    y_offset = -0.8  # 放在时间线下方，有更大的距离
                
                # 4. 在"满花"的情况下：将"代用样品发送"放到时间线上方
                if dept == "产前确认" and step == "代用样品发送" and (process_type == "满花" or process_type == "局花" or process_type == "绣花"):
                    y_offset = 0.3  # 放在时间线上方
                
                text = ax.text(text_x, y + y_offset, step_text, 
                               ha='center', 
                               va='bottom' if y_offset > 0 else 'top',
                               fontsize=16, 
                               color='black', 
                               fontweight='bold',
                               bbox=text_box,
                               zorder=5,  # Ensure text is above other elements
                               snap=True,
                               fontproperties=prop)  # Snap to pixel grid
        
        # 绘制实线连接
        x_positions = sorted(list(time_groups.keys()))
        if len(x_positions) > 1:
            ax.plot(x_positions, [y] * len(x_positions), '-', 
                   color='black', 
                   alpha=0.7, 
                   zorder=2,
                   linewidth=1.5,
                   solid_capstyle='round',
                   snap=True)  # Snap to pixel grid
    
    # # 在右侧添加缝纫部分
    # 缝纫_steps = schedule["缝纫"]
    # y_center = (max(y_positions.values()) + min(y_positions.values())) / 2
    # ax.fill_betweenx([min(y_positions.values()) - 0.4, max(y_positions.values()) + 0.4], 
    #                  0.92, 1.0, color=department_colors["缝纫"], alpha=0.5)
    
    # # 绘制缝纫步骤
    # step_names = list(缝纫_steps.keys())
    # step_dates = [缝纫_steps[step]["时间点"] for step in step_names]
    
    # # 按时间顺序排序步骤
    # sorted_steps = sorted(zip(step_names, step_dates), key=lambda x: x[1])
    
    # # 计算x轴位置（在0.94-0.99之间平均分布）
    # x_positions = []
    # if len(sorted_steps) > 1:
    #     x_start = 0.93
    #     x_end = 0.99
    #     x_step = (x_end - x_start) / (len(sorted_steps) - 1)
    #     x_positions = [x_start + i * x_step for i in range(len(sorted_steps))]
    # else:
    #     x_positions = [0.955]  # 如果只有一个步骤，放在中间
    
    # # 绘制步骤和连接线
    # for i, ((step, date), x_pos) in enumerate(zip(sorted_steps, x_positions)):
    #     # 绘制点
    #     ax.scatter(x_pos, y_center, color='black', zorder=3)
        
    #     # 为缝纫步骤添加文本框
    #     text_box = dict(boxstyle='round,pad=0.4', facecolor='white', alpha=0.8, edgecolor='black', linewidth=1)
    #     step_text = f"{step}\n{date.strftime('%Y/%m/%d')}"
    #     ax.text(x_pos, y_center - 0.3, step_text, ha='center', va='top',
    #            fontsize=16, color='black', fontweight='bold',
    #            bbox=text_box, fontproperties=prop)
    
    ## 绘制连接线
    #if len(x_positions) > 1:
    #    ax.plot(x_positions, [y_center] * len(x_positions), '-', color='black', alpha=0.7, zorder=2)
    
    # 设置坐标轴
    ax.set_yticks(list(y_positions.values()))
    ax.set_yticklabels(list(y_positions.keys()), fontsize=22, fontweight='bold', fontproperties=prop)  # 统一部门标签大小
    ax.set_xticks([])
    ax.set_xticklabels([])
    ax.set_xlim(-0.02, 1.02)
    ax.set_ylim(min(y_positions.values()) - 0.7, max(y_positions.values()) + 0.7)  # 减小y轴的上下边距
    
    # 设置标题
    title_text = f"生产流程时间表 - {process_type}"
    if "style_number" in st.session_state and st.session_state["style_number"]:
        style_number_text = "款号: " + str(st.session_state["style_number"])
        # 将款号分成每行最多30个字符
        style_number_wrapped = [style_number_text[i:i+30] for i in range(0, len(style_number_text), 30)]
        # Add production group if available
        if "production_group" in st.session_state and st.session_state["production_group"]:
            group_text = f"生产组: {st.session_state['production_group']}"
            style_number_wrapped.append(group_text)
        title_text += "\n" + "\n".join(style_number_wrapped)
    ax.set_title(title_text, fontsize=30, fontweight='bold', y=1.02 + 0.02 * (len(style_number_wrapped) if 'style_number_wrapped' in locals() else 0), fontproperties=prop)
    ax.set_frame_on(False)
    
    # 调整图形布局以适应文本框
    plt.subplots_adjust(left=0.1, right=0.9, bottom=0.1, top=0.98)
    
    return fig  # Return the figure instead of displaying it

# Function to generate department-specific plots
def generate_department_wise_plots(styles):
    all_schedules = []
    department_colors = {
            "产前确认": "#FFF0C1",
            "面料": "#FFDDC1", 
            "满花": "#C1E1FF", 
            "裁剪": "#D1FFC1", 
            "局花": "#FFC1E1", 
            "绣花": "#E1C1FF", 
            "配片": "#FFD1C1", 
            "滚领": "#C1FFD1", 
            "辅料": "#E1FFC1", 
            "缝纫": "#FFC1C1",
            "后整": "#FFE1C1",
            "工艺": "#C1FFE1"
        }
    
    # Calculate schedules for all styles
    for style in styles:
        sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
        start_time_period = style.get("start_time_period", "上午")  # 获取上午/下午信息
        schedule = calculate_schedule(
            sewing_start_time, 
            style["process_type"], 
            style["cycle"], 
            style["order_quantity"], 
            style["daily_production"],
            start_time_period
        )
        for dept, steps in schedule.items():
            for step, data in steps.items():
                # 创建一个新字典来存储步骤数据
                step_data = {
                    "style_number": style["style_number"],
                    "department": dept,
                    "step": step,
                    "date": data["时间点"],
                    "process_type": style["process_type"],
                    "production_group": style.get("production_group", "")
                }
                
                # 如果有备注信息，添加到步骤数据中
                if "备注" in data:
                    step_data["remarks"] = data["备注"]
                
                all_schedules.append(step_data)
    
    # Convert to DataFrame for sorting
    df = pd.DataFrame(all_schedules)
    
    # Create a temporary directory
    temp_dir = tempfile.mkdtemp()
    
    # Generate department-wise plots
    for department in df["department"].unique():
        dept_data = df[df["department"] == department].copy()
        
        # Sort by date (latest first) and then by style number
       # 1. 计算每个款式的最早步骤时间
        earliest_dates = dept_data.groupby("style_number")["date"].min().reset_index()
        earliest_dates.rename(columns={"date": "earliest_date"}, inplace=True)
        
        # 2. 将最早日期合并回原始数据
        dept_data = pd.merge(dept_data, earliest_dates, on="style_number", how="left")
        
        # 3. 按最早日期排序（升序，最早的在前），然后按款式号排序
        dept_data.sort_values(by=["earliest_date", "style_number"], ascending=[False, False], inplace=True)
        
        # 4. 获取排序后的唯一款式号列表（保持顺序）
        unique_sorted_styles = dept_data["style_number"].unique()
        
        # Calculate time range for dynamic sizing
        date_range = (dept_data["date"].max() - dept_data["date"].min()).days
        base_width = int(date_range/41*40)
        if base_width > 24:
            dpi_scale = base_width / 24
            plt.rcParams['figure.dpi'] = int(300 * dpi_scale)
            plt.rcParams['savefig.dpi'] = int(300 * dpi_scale)
        
        # Create figure with dynamic sizing
        fig, ax = plt.subplots(figsize=(max(base_width, 25), len(unique_sorted_styles) * 3))
        fig.patch.set_facecolor('white')
        ax.set_facecolor('white')
        
        # Calculate y positions for each style number - use the sorted styles
        y_positions = {style: idx * 1.5 for idx, style in enumerate(unique_sorted_styles)}
        
        # Create colored background for the department
        ax.fill_betweenx(
            [min(y_positions.values()) - 0.4, max(y_positions.values()) + 0.4],
            0, 1,
            color=department_colors.get(department, "#DDDDDD"),
            alpha=0.5
        )
        
        # Plot timeline for each style number
        for style in unique_sorted_styles:
            style_data = dept_data[dept_data["style_number"] == style]
            y = y_positions[style]
            
            # Convert dates to relative positions (0 to 1)
            date_min = dept_data["date"].min()
            date_max = dept_data["date"].max()
            total_days = (date_max - date_min).days
            
            # Sort steps by date
            style_data = style_data.sort_values(by="date")
            
            # Group steps by date
            date_groups = {}
            for _, row in style_data.iterrows():
                date_key = row["date"]
                if date_key not in date_groups:
                    date_groups[date_key] = []
                date_groups[date_key].append(row)
            
            # Plot points and labels for each date group
            x_positions = []
            dates = list(date_groups.keys())
            
            for date_idx, (date, rows) in enumerate(date_groups.items()):
                # Calculate x position
                if total_days == 0:
                    x_pos = 0.5  # Center of the timeline
                else:
                    days_from_start = (date - date_min).days
                    x_pos = 0.1 + (days_from_start / total_days) * 0.8  # Leave margins
                x_positions.append(x_pos)
                
                # Plot point
                ax.scatter(x_pos, y, color='black', zorder=3)
                
                # Calculate text position based on adjacent dates
                text_x = x_pos
                
                # Check if there's a previous or next date within 1 day
                prev_date = dates[date_idx-1] if date_idx > 0 else None
                next_date = dates[date_idx+1] if date_idx < len(dates)-1 else None
                
                scaling_factor = 0.015 + (0.04 * (1 - min(1, total_days / 20)))  # ✅ Adjust dynamically
                # Adjust text position if dates are 1 day apart
                if prev_date and abs((date - prev_date).days) == 1:
                    text_x = x_pos + scaling_factor#0.015  # Move right
                elif next_date and abs((date - next_date).days) == 1:
                    text_x = x_pos - scaling_factor#0.015  # Move left
                
                # Stack text boxes for steps on the same day
                for i, row in enumerate(rows):
                    text_box = dict(
                        boxstyle='round,pad=0.4',
                        facecolor='white',
                        alpha=1.0,
                        edgecolor='black',
                        linewidth=1
                    )
                    
                    # Calculate vertical offset for stacking
                    y_offset = -0.3 - i * 0.3  # Stack boxes vertically
                    
                    # Special handling for 产前确认, 面料, place it above the timeline
                    if ((department == "产前确认" and (row["step"] == "色样确认" or row["step"] == "绣花样品"))
                        or (department == "面料" and row["step"] == "工艺分析")
                        or (department == "满花" and row["step"] == "满花后整")
                        or (department == "后整" and row["step"] == "包装")):
                        y_offset = 0.3  # Place above the timeline

                    # 1. 对于包含"满花"的流程（除了"满花绣花"）：将"满花样品"放到时间线上方
                    if department == "产前确认" and row["step"] == "满花样品":
                        # 查找样式信息以获取流程类型
                        for style_info in styles:
                            if style_info["style_number"] == row["style_number"]:
                                process_type = style_info.get("process_type", "")
                                if "满花" in process_type and process_type != "满花绣花":
                                    y_offset = 0.3  # 放在时间线上方
                                #break
                    
                    # 2. 在"满花局花绣花"的情况下：将"局花样品"放到时间线上方，并与时间线保持一个文本框的距离
                    if department == "产前确认" and row["step"] == "局花样品":
                        # 查找样式信息以获取流程类型
                        for style_info in styles:
                            if style_info["style_number"] == row["style_number"]:
                                process_type = style_info.get("process_type", "")
                                if process_type == "满花局花绣花":
                                    y_offset = 0.6  # 放在时间线上方，有更大的距离
                                #break
                    
                    # 3. 除了"满花局花绣花"或"满花"的情况下：将"版型"步骤放到时间线下方，与时间线有一个文本框的距离
                    if department == "产前确认" and row["step"] == "版型":
                        # 查找样式信息以获取流程类型
                        for style_info in styles:
                            if style_info["style_number"] == row["style_number"]:
                                process_type = style_info.get("process_type", "")
                                if process_type != "满花局花绣花" and process_type != "满花" and process_type != "局花绣花" and process_type != "绣花":
                                    y_offset = -0.6  # 放在时间线下方，有更大的距离
                                #break
                    
                    # 4. 在"满花"的情况下：将"代用样品发送"放到时间线下方
                    if department == "产前确认" and row["step"] == "代用样品发送":
                        # 查找样式信息以获取流程类型
                        for style_info in styles:
                            if style_info["style_number"] == row["style_number"]:
                                process_type = style_info.get("process_type", "")
                                if process_type == "满花":
                                    y_offset = -0.6  # 放在时间线下方
                                elif process_type == "局花" or process_type == "绣花":
                                    y_offset = 0.3  # 放在时间线上方
                                #break

                                
                    step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}"
                    
                    # 为部门时间线图的单独绘制中添加备注显示
                    # 查找原始数据中的备注信息 - 使用缓存的schedule数据
                    if department == "缝纫" and (row["step"] == "缝纫结束" or row["step"] == "缝纫开始"):
                         # DataFrame行访问需要用不同的方式
                        if "remarks" in row and pd.notna(row["remarks"]) and row["remarks"]:
                            # 如果DataFrame行中有remarks数据
                            step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{row['remarks']}"
                        else:
                            # 作为备份，从原始style数据中查找
                            for style_info in styles:
                                if style_info["style_number"] == row["style_number"]:
                                    if row["step"] == "缝纫开始":
                                        start_time_period = style_info.get("start_time_period", "上午")
                                        step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{start_time_period}"
                                    elif row["step"] == "缝纫结束" and "schedule" in style_info:
                                        if "缝纫" in style_info["schedule"] and "缝纫结束" in style_info["schedule"]["缝纫"] and "备注" in style_info["schedule"]["缝纫"]["缝纫结束"]:
                                            end_remark = style_info["schedule"]["缝纫"]["缝纫结束"]["备注"]
                                            step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{end_remark}"
                                    break
                    
                    ax.text(
                        text_x, y + y_offset,
                        step_text,
                        ha='center',
                        va='bottom' if y_offset > 0 else 'top',  # Adjust vertical alignment based on position
                        fontsize=12,
                        fontweight='bold',
                        bbox=text_box,
                        zorder=5, fontproperties=prop
                    )
            
            # Connect points with lines
            if len(x_positions) > 1:
                ax.plot(x_positions, [y] * len(x_positions), '-',
                       color='black',
                       alpha=0.7,
                       zorder=2,
                       linewidth=1.5)
        
        # Set up the axes
        ax.set_yticks(list(y_positions.values()))
        # Include production group in y-axis labels if available
        y_labels = []
        for style in y_positions.keys():
            style_rows = dept_data[dept_data["style_number"] == style]
            production_group = style_rows.iloc[0]["production_group"] if len(style_rows) > 0 and style_rows.iloc[0]["production_group"] else ""
            # 查找生产顺序
            original_style = next((s for s in styles if s["style_number"] == style), None)
            if original_style and "production_order" in original_style:
                production_order = original_style["production_order"]
                if production_group:
                    y_labels.append(f"款号: {style} (生产组: {production_group}, 序号: {production_order})")
                else:
                    y_labels.append(f"款号: {style} (序号: {production_order})")
            else:
                if production_group:
                    y_labels.append(f"款号: {style} (生产组: {production_group})")
                else:
                    y_labels.append(f"款号: {style}")
        
        ax.set_yticklabels(y_labels, fontsize=14, fontweight='bold', fontproperties=prop)
        ax.set_xticks([])
        ax.set_xlim(-0.02, 1.02)
        ax.set_ylim(min(y_positions.values()) - 0.7, max(y_positions.values()) + 0.7)
        
        # Set title
        ax.set_title(department,
                    fontsize=24,
                    fontweight='bold',
                    y=1.02, fontproperties=prop)
        ax.set_frame_on(False)
        
        # Save figure
        fig_path = os.path.join(temp_dir, f"{department}.png")
        fig.savefig(fig_path, dpi=300, bbox_inches="tight")
        plt.close(fig)
    
    # Now create production group specific plots - only for 缝纫 department
    for department in df["department"].unique():
        # Skip all departments except 缝纫
        if department != "缝纫":
            continue
            
        # Get unique production groups for this department
        dept_data = df[df["department"] == department].copy()
        production_groups = dept_data["production_group"].unique()
        
        for group in production_groups:
            if not group:  # Skip empty production groups
                continue
                
            # Filter data for this production group
            group_data = dept_data[dept_data["production_group"] == group].copy()
            
            # If we don't have enough data, skip
            if len(group_data) == 0 or len(group_data["style_number"].unique()) == 0:
                continue

            # 为生产组图表也应用相似的排序逻辑
            # 1. 计算每个款式的最早步骤时间
            earliest_dates = group_data.groupby("style_number")["date"].min().reset_index()
            earliest_dates.rename(columns={"date": "earliest_date"}, inplace=True)
            
            # 2. 将最早日期合并回原始数据
            group_data = pd.merge(group_data, earliest_dates, on="style_number", how="left")
            
            # 3. 按最早日期排序（升序，最早的在前），然后按款式号排序
            group_data.sort_values(by=["earliest_date", "style_number"], ascending=[False, False], inplace=True)
            
            # 4. 获取排序后的唯一款式号列表（保持顺序）
            unique_sorted_styles = group_data["style_number"].unique()
            
            # Create figure
            base_width = max(20, int((group_data["date"].max() - group_data["date"].min()).days / 41 * 40))
            fig, ax = plt.subplots(figsize=(base_width, len(unique_sorted_styles) * 3))
            fig.patch.set_facecolor('white')
            ax.set_facecolor('white')
            
            y_positions = {style: i for i, style in enumerate(unique_sorted_styles)}
            
            # Plot timeline for each style
            for style, y in y_positions.items():
                style_data = group_data[group_data["style_number"] == style].sort_values("date")
                
                # Normalize dates to 0-1 range for x-axis
                date_range = (group_data["date"].max() - group_data["date"].min()).days
                if date_range == 0:
                    date_range = 1  # Avoid division by zero
                
                min_date = group_data["date"].min()
                
                # Draw points and text for each step
                x_positions = []
                
                for _, row in style_data.iterrows():
                    # Calculate normalized position on x-axis
                    x = (row["date"] - min_date).days / date_range
                    x_positions.append(x)
                    
                    # Draw point - using standard style
                    ax.scatter(x, y, s=100, color='blue', edgecolor='black', zorder=3)
                    
                    # Add text with step name and date
                    # Adjust position based on step type
                    text_x = x
                    y_offset = -0.3  # Default to below the timeline
                    
                    # Special text box for certain steps
                    text_box = dict(boxstyle="round,pad=0.3", facecolor='lightyellow', alpha=0.7, edgecolor='black')
                    
                    # Change position for certain steps
                    if ((department == "裁床" and row["step"] == "裁剪完成") or 
                        (department == "缝纫" and (row["step"] == "缝纫结束" or row["step"] == "缝纫开始")) or 
                        (department == "后整" and row["step"] == "包装")):
                        y_offset = 0.3  # Place above the timeline
                    
                    step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}"
                    
                    # 为部门时间线图的单独绘制中添加备注显示
                    # 查找原始数据中的备注信息 - 使用缓存的schedule数据
                    if department == "缝纫" and (row["step"] == "缝纫结束" or row["step"] == "缝纫开始"):
                        # DataFrame行访问需要用不同的方式
                        if "remarks" in row and pd.notna(row["remarks"]) and row["remarks"]:
                            # 如果DataFrame行中有remarks数据
                            step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{row['remarks']}"
                        else:
                            # 作为备份，从原始style数据中查找
                            for style_info in styles:
                                if style_info["style_number"] == row["style_number"]:
                                    if row["step"] == "缝纫开始":
                                        start_time_period = style_info.get("start_time_period", "上午")
                                        step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{start_time_period}"
                                    elif row["step"] == "缝纫结束" and "schedule" in style_info:
                                        if "缝纫" in style_info["schedule"] and "缝纫结束" in style_info["schedule"]["缝纫"] and "备注" in style_info["schedule"]["缝纫"]["缝纫结束"]:
                                            end_remark = style_info["schedule"]["缝纫"]["缝纫结束"]["备注"]
                                            step_text = f"{row['step']}\n{row['date'].strftime('%Y/%m/%d')}\n{end_remark}"
                                    break
                    ax.text(
                        text_x, y + y_offset,
                        step_text,
                        ha='center',
                        va='bottom' if y_offset > 0 else 'top',  # Adjust vertical alignment based on position
                        fontsize=12,
                        fontweight='bold',
                        bbox=text_box,
                        zorder=5, fontproperties=prop
                    )
                
                # Connect points with lines
                if len(x_positions) > 1:
                    ax.plot(x_positions, [y] * len(x_positions), '-',
                           color='black',
                           alpha=0.7,
                           zorder=2,
                           linewidth=1.5)
            
            # Set up the axes
            ax.set_yticks(list(y_positions.values()))
            # Include production group in y-axis labels if available
            y_labels = []
            for style in y_positions.keys():
                style_rows = group_data[group_data["style_number"] == style]
                production_group = style_rows.iloc[0]["production_group"] if len(style_rows) > 0 and style_rows.iloc[0]["production_group"] else ""
                
                # 查找生产顺序
                original_style = next((s for s in styles if s["style_number"] == style), None)
                if original_style and "production_order" in original_style:
                    production_order = original_style["production_order"]
                    if production_group:
                        y_labels.append(f"款号: {style} (生产组: {production_group}, 序号: {production_order})")
                    else:
                        y_labels.append(f"款号: {style} (序号: {production_order})")
                else:
                    if production_group:
                        y_labels.append(f"款号: {style} (生产组: {production_group})")
                    else:
                        y_labels.append(f"款号: {style}")
            
            ax.set_yticklabels(y_labels, fontsize=14, fontweight='bold', fontproperties=prop)
            ax.set_xticks([])
            ax.set_xlim(-0.02, 1.02)
            ax.set_ylim(min(y_positions.values()) - 0.7, max(y_positions.values()) + 0.7)
            
            # Set title to include production group - using standard style
            ax.set_title(f"{department} - 生产组: {group}",
                        fontsize=24,
                        fontweight='bold',
                        y=1.02, fontproperties=prop)
            ax.set_frame_on(False)

            # Save with production group in filename
            group_fig_path = os.path.join(temp_dir, f"{department}_生产组_{group}.png")
            fig.savefig(group_fig_path, dpi=300, bbox_inches="tight")
            plt.close(fig)
    
    # Create ZIP archive
    zip_path = os.path.join(temp_dir, "Department_Timelines.zip")
    with zipfile.ZipFile(zip_path, 'w') as zipf:
        for file in os.listdir(temp_dir):
            if file.endswith(".png"):
                zipf.write(
                    os.path.join(temp_dir, file),
                    file  # Just use the filename directly
                )
    
    return zip_path
    
def adjust_schedule(schedule, department, delayed_step, new_end_time):
    if department not in schedule or delayed_step not in schedule[department]:
        return schedule
    
    delay_days = (new_end_time - schedule[department][delayed_step]["时间点"]).days
    found_delayed_step = False
    
    for step, times in schedule[department].items():
        if step == delayed_step:
            found_delayed_step = True
            schedule[department][step]["时间点"] = new_end_time
        elif found_delayed_step:
            schedule[department][step]["时间点"] += timedelta(days=delay_days)
    
    return schedule 

# Define valid credentials (you can modify this dictionary as needed)
VALID_CREDENTIALS = {
    "admin": "JD2024",
    "user1": "password1",
    "user2": "password2"
}

# Initialize session state for login
if "logged_in" not in st.session_state:
    st.session_state["logged_in"] = False
if "current_user" not in st.session_state:
    st.session_state["current_user"] = None

def login(account_id):
    st.session_state["logged_in"] = True
    st.session_state["current_user"] = account_id
    # Load user's saved data
    user_data = load_user_data(account_id)
    st.session_state["all_styles"] = user_data["all_styles"]

# Login page


if not st.session_state.get("logged_in", False):
    # ✅ 使用 st.empty() 确保所有内容填充整个页面
    login_container = st.empty()

    with login_container.container():
        # ✅ 创建两列布局
        col1, col2 = st.columns([1, 1])  # 左侧登录，右侧欢迎信息

        # 🎨 **左侧：登录框**
        with col1:
            st.markdown(
                """
                <div style="min-width: 500px; max-width: 700px; 
                            padding: 40px;  /* ✅ 让整个左边框更美观 */
                            background-color: white; 
                            border-radius: 10px;">
                    <h2 style='text-align: left; 
                            margin-top: 40px;  
                            margin-bottom: 10px;  
                            font-size: 2.5em;'>
                        登录到您的账户
                    </h2>
                </div>
                """,
                unsafe_allow_html=True
            )
            # 如果有LOGO，可以放在这里
            # st.image("logo.png", width=120)

            account_id = st.text_input("账号", key="account_input")
            password = st.text_input("密码", type="password", key="password_input")

            # ✅ 居中的登录按钮
            col_a, col_b, col_c = st.columns([1, 2, 1])
            with col_b:
                button_style = """
                <style>
                    div[data-testid="stButton"] button {
                        background: linear-gradient(135deg, #6a11cb, #2575fc);
                        color: white;
                        border: none;
                        padding: 0.5rem 1rem;
                        border-radius: 5px;
                        font-weight: bold;
                        transition: all 0.3s ease;
                    }
                    div[data-testid="stButton"] button:hover {
                        background: linear-gradient(135deg, #5a0cb1, #1e63d6);
                        transform: translateY(-2px);
                        box-shadow: 0 4px 12px rgba(106, 17, 203, 0.3);
                    }
                </style>
                """
                st.markdown(button_style, unsafe_allow_html=True)
                if st.button("登录", use_container_width=True):
                    if account_id in VALID_CREDENTIALS and password == VALID_CREDENTIALS[account_id]:
                        st.session_state["logged_in"] = True
                        st.session_state["current_user"] = account_id
                        st.rerun()
                    else:
                        st.error("账号或密码错误，请重试")

        # 🎨 **右侧：欢迎信息**
        with col2:
            st.markdown(
                """
                <div style='
                    background: linear-gradient(135deg, #6a11cb, #2575fc);
                    padding: 70px;
                    min-height: 500px;
                    min-width: 450px;
                    color: white;
                    border-radius: 120px 40px 40px 120px;  /* ✅ 让它更符合UI */
                    text-align: center;
                    margin-left: 100px;
                '>
                    <h1 style="margin-bottom: 10px; color: white;">欢迎回来！</h1>
                    <p style="font-size: 18px; color: white;">请登录以访问生产流程管理系统。</p>
                </div>
                """,
                unsafe_allow_html=True
            )

    # ✅ 强制设置高度，防止需要滚动
    st.markdown(
        """
        <style>
        .block-container {
            padding-top: 3vh !important;  /* 页面上方留空间 */
            height: 90vh !important;  /* 让整个界面占满 */
            max-width: 1600px !important; /* 控制最大宽度 */
            display: flex;
            justify-content: center;
            align-items: center;
        }
        </style>
        """,
        unsafe_allow_html=True
    )


else:
    # Main application code
    st.title("生产流程时间管理系统")
    
    # Add user info and logout button in the top right
    col1, col2, col3 = st.columns([8, 2, 1])
    with col2:
        st.write(f"当前用户: {st.session_state['current_user']}")
    with col3:
        if st.button("登出"):
            # Save user data before logging out
            save_user_data(st.session_state["current_user"], {
                "all_styles": st.session_state["all_styles"]
            })
            st.session_state["logged_in"] = False
            st.session_state["current_user"] = None
            st.rerun()
    
    # Initialize session state
    if "all_styles" not in st.session_state:
        st.session_state["all_styles"] = []

    # 添加Excel上传功能
    st.subheader("方式一：上传Excel文件")
    uploaded_file = st.file_uploader("上传Excel文件 (必需列：款号、缝纫开始日期、缝纫开始时间、工序、确认周转周期, '订单数量', '日产量', '生产组', '生产顺序')", type=['xlsx', 'xls'])


    if uploaded_file is not None:
        try:
            df = pd.read_excel(uploaded_file)
            required_columns = ['款号', '缝纫开始日期', '缝纫开始时间', '工序', '确认周转周期', '订单数量', '日产量', '生产组', '生产顺序']
            # 显示Excel可选列的说明
            st.info("""
            **Excel文件说明**:
            * 必需列: 款号、缝纫开始日期、缝纫开始时间、工序、确认周转周期、订单数量、日产量、生产组
            * 可选列: 生产顺序 (同一生产组内款式的排产顺序，相同顺序号的款式将在同一天开始生产)
            * 缝纫开始时间列应填写"上午"或"下午"
            * 工序列应为以下之一: 满花局花绣花、满花局花、满花绣花、局花绣花、满花、局花、绣花
            """)
            
            # Check if all required columns exist
            if not all(col in df.columns for col in required_columns):
                st.error(f"Excel文件必须包含以下列：{', '.join(required_columns)}")
            else:
                # Convert dates to datetime if they aren't already
                df['缝纫开始日期'] = pd.to_datetime(df['缝纫开始日期']).dt.date
                
                # Validate process types
                valid_processes = ["满花局花绣花", "满花局花", "满花绣花", "局花绣花", "满花", "局花", "绣花"]
                invalid_processes = df[~df['工序'].isin(valid_processes)]['工序'].unique()
                if len(invalid_processes) > 0:
                    st.error(f"发现无效的工序类型：{', '.join(invalid_processes)}")
                else:
                    # 检查"生产顺序"列是否存在
                    has_production_order = '生产顺序' in df.columns
                    if has_production_order:
                        st.success("检测到'生产顺序'列，将根据此列对同一生产组内的款式进行排序。")
                        
                    # Add new styles from Excel
                    new_styles = []
                    for _, row in df.iterrows():
                        # 确保缝纫开始时间是上午或下午，默认为上午
                        start_time = row['缝纫开始时间'] if row['缝纫开始时间'] in ["上午", "下午"] else "上午"
                        # 获取生产顺序，如果存在
                        production_order = int(row['生产顺序']) if '生产顺序' in df.columns and pd.notna(row['生产顺序']) else 1
                        
                        new_style = {
                            "style_number": str(row['款号']),
                            "sewing_start_date": row['缝纫开始日期'],
                            "start_time_period": start_time,
                            "process_type": row['工序'],
                            "cycle": int(row['确认周转周期']),
                            "order_quantity": int(row['订单数量']),
                            "daily_production": int(row['日产量']),
                            "production_group": str(row['生产组']),
                            "production_order": production_order
                        }
                        new_styles.append(new_style)
                    
                    if st.button("添加Excel中的款号"):
                        st.session_state["all_styles"].extend(new_styles)
                        # Auto-save after adding styles
                        save_user_data(st.session_state["current_user"], {
                            "all_styles": st.session_state["all_styles"]
                        })
                        st.success(f"已从Excel添加 {len(new_styles)} 个款号")
                        st.rerun()
        
        except Exception as e:
            st.error(f"读取Excel文件时出错：{str(e)}")

    st.subheader("方式二：手动输入")
    # 创建输入表单
    with st.form("style_input_form"):
        # 批量输入款号，每行一个
        style_numbers = st.text_area("请输入款号(每行一个):", "")
        sewing_start_date = st.date_input("请选择缝纫开始日期:", min_value=datetime.today().date())

        # 添加上午/下午选择
        col1, col2 = st.columns(2)
        with col1:
            start_time_period = st.selectbox("缝纫开始时间:", ["上午", "下午"])
        with col2:
            process_options = ["满花局花绣花", "满花局花", "满花绣花", "局花绣花", "满花", "局花", "绣花"]
            selected_process = st.selectbox("请选择工序:", process_options)
        # 添加新字段
        order_quantity = st.number_input("订单数量:", min_value=1, value=100)
        daily_production = st.number_input("日产量:", min_value=1, value=50)
        production_group = st.text_input("生产组号:", "")
        production_order = st.number_input("生产顺序:", min_value=1, value=1, help="同一生产组内款式的生产顺序")
        cycle = st.selectbox("请选择确认周转周期:", [7, 14, 30])
        
        submitted = st.form_submit_button("添加款号")
        if submitted and style_numbers:
            # 分割多行输入，去除空行和空格
            new_style_numbers = [s.strip() for s in style_numbers.split('\n') if s.strip()]
            
            # 添加新的款号信息
            for style_number in new_style_numbers:
                new_style = {
                    "style_number": style_number,
                    "sewing_start_date": sewing_start_date,
                    "start_time_period": start_time_period,  # 新增上午/下午字段
                    "process_type": selected_process,
                    "cycle": cycle,
                    "order_quantity": order_quantity,
                    "daily_production": daily_production,
                    "production_group": production_group,
                    "production_order": production_order
                }
                st.session_state["all_styles"].append(new_style)
            # Auto-save after adding styles
            save_user_data(st.session_state["current_user"], {
                "all_styles": st.session_state["all_styles"]
            })
            st.success(f"已添加 {len(new_style_numbers)} 个款号")

    # 显示当前添加的所有款号
    if st.session_state["all_styles"]:
        st.subheader("已添加的款号:")
        
        # 使用列表来显示所有款号，并提供删除按钮
        for idx, style in enumerate(st.session_state["all_styles"]):
            col1, col2 = st.columns([4, 1])
            with col1:
                time_period = style.get("start_time_period", "上午")  # 默认为上午
                production_order = style.get("production_order", "-")
                st.write(f"{idx + 1}. 款号: {style['style_number']}, 工序: {style['process_type']}, " 
                    f"缝纫开始日期: {style['sewing_start_date']} {time_period}, 周期: {style['cycle']}, "
                    f"订单数量: {style.get('order_quantity', '-')}, 日产量: {style.get('daily_production', '-')}, "
                    f"生产组号: {style.get('production_group', '-')}, 生产顺序: {production_order}")
            with col2:
                if st.button("删除", key=f"delete_{idx}"):
                    st.session_state["all_styles"].pop(idx)
                    # Auto-save after deleting style
                    save_user_data(st.session_state["current_user"], {
                        "all_styles": st.session_state["all_styles"]
                    })
                    st.rerun()
        
        # 添加清空所有按钮
        if st.button("清空所有款号"):
            st.session_state["all_styles"] = []
            # Auto-save after clearing styles
            save_user_data(st.session_state["current_user"], {
                "all_styles": st.session_state["all_styles"]
            })
            st.rerun()

    # 添加是否启用连续排产的选项
    if st.session_state["all_styles"]:
        st.subheader("生成图表")

        # 添加连续排产逻辑说明
        with st.expander("📋 查看生产组连续排产逻辑说明"):
            st.markdown("""
            ### 生产组连续排产逻辑
            
            系统按以下规则处理同一生产组内的款式排产:
            
            1. **同一生产顺序的款式**:
               - 共享相同的缝纫开始日期和时段（上午/下午）
               - 生产顺序为1的款式使用原始设定的开始日期
               - 各自按其工序、订单数量和日产量计算结束时间
            
            2. **连续排产规则**:
               - 系统会找出当前生产顺序组中结束时间最晚的款式
               - 该款式的缝纫结束时间（及上午/下午时段）将作为下一个生产顺序组的开始时间
               - 依此类推，形成连续排产
               
            3. **实际应用**:
               - 生产顺序为1的款式可以手动指定开始日期
               - 生产顺序为2、3...的款式会自动根据前一组的结束时间进行排产
               - 相同生产顺序的款式将在同一天同一时段开始，可能在不同时间结束
            """)
        
        enable_sequential_production = st.checkbox("启用生产组连续排产功能", value=True, 
                                            help="启用后，同一生产组内，下一个生产顺序(production_order)的款式将从前一个生产顺序中最晚完成的款式结束时间开始")
        
        # 添加预览按钮
        if enable_sequential_production and st.button("预览生产组排产结果"):
            # 重新安排同一生产组内款式的缝纫开始时间
            preview_styles = rearrange_styles_by_production_group(st.session_state["all_styles"])
            
            # 按生产组分组显示排产结果
            grouped_styles = {}
            for style in preview_styles:
                group = style.get("production_group", "无生产组")
                if group not in grouped_styles:
                    grouped_styles[group] = []
                grouped_styles[group].append(style)
            
            # 显示每个生产组的排产结果
            for group, styles in grouped_styles.items():
                if group != "无生产组":
                    st.write(f"### 生产组: {group}")
                    
                    # 按生产顺序进一步分组
                    order_grouped_styles = {}
                    for style in styles:
                        order = style.get("production_order", 9999)
                        if order not in order_grouped_styles:
                            order_grouped_styles[order] = []
                        order_grouped_styles[order].append(style)
                    
                    # 记录前一个顺序组的结束信息，用于显示连续关系
                    prev_end_info = None
                    
                    # 按生产顺序排序
                    for order in sorted(order_grouped_styles.keys()):
                        order_styles = order_grouped_styles[order]
                        
                        # 如果不是第一个生产顺序，显示连续关系
                        if prev_end_info:
                            st.markdown(f"""
                            <div style="text-align:center; padding: 10px; margin: 15px 0; background-color: #f0f2f6; border-radius: 5px;">
                                ⬇️ <b>前一个生产顺序组最晚完成的款式 {prev_end_info['style']} 
                                结束时间: {prev_end_info['date']} ({prev_end_info['remark']})</b>
                            </div>
                            """, unsafe_allow_html=True)
                        
                        st.write(f"#### 生产顺序: {order}")
                        
                        # 创建数据表
                        preview_data = []
                        for style in order_styles:
                            # 计算缝纫结束时间
                            sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
                            start_time_period = style.get("start_time_period", "上午")
                            schedule = calculate_schedule(
                                sewing_start_time, 
                                style["process_type"], 
                                style["cycle"], 
                                style["order_quantity"], 
                                style["daily_production"],
                                start_time_period
                            )
                            
                            sewing_end_time = schedule["缝纫"]["缝纫结束"]["时间点"]
                            sewing_end_remark = schedule["缝纫"]["缝纫结束"].get("备注", "")
                            
                            preview_data.append({
                                "款号": style["style_number"],
                                "工序": style["process_type"],
                                "缝纫开始日期": f"{style['sewing_start_date']} ({start_time_period})",
                                "缝纫结束日期": f"{sewing_end_time.date()} ({sewing_end_remark})",
                                "订单数量": style["order_quantity"],
                                "日产量": style["daily_production"],
                                "生产天数": round(style["order_quantity"] * 1.05 / style["daily_production"], 1)
                            })
                        
                        # 显示表格
                        st.table(preview_data)
                        
                        # 如果这个组有多个款式，计算并显示组内最晚结束时间
                        latest_end_time = None
                        latest_end_remark = ""
                        latest_style = None
                        
                        for style in order_styles:
                            sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
                            start_time_period = style.get("start_time_period", "上午")
                            schedule = calculate_schedule(
                                sewing_start_time, 
                                style["process_type"], 
                                style["cycle"], 
                                style["order_quantity"], 
                                style["daily_production"],
                                start_time_period
                            )
                            
                            end_time = schedule["缝纫"]["缝纫结束"]["时间点"]
                            end_remark = schedule["缝纫"]["缝纫结束"].get("备注", "")
                            
                            if latest_end_time is None or end_time > latest_end_time:
                                latest_end_time = end_time
                                latest_end_remark = end_remark
                                latest_style = style["style_number"]
                        
                        # 更新前一个顺序组的结束信息，用于下一个顺序组的显示
                        prev_end_info = {
                            "style": latest_style,
                            "date": latest_end_time.date(),
                            "remark": latest_end_remark
                        }
                        
                        if len(order_styles) > 1:
                            st.info(f"⚠️ 注意：该生产顺序组中，款号 **{latest_style}** 的缝纫结束时间最晚：**{latest_end_time.date()} ({latest_end_remark})**，下一个生产顺序组将从此时间开始。")
            
            # 显示无生产组的款式
            if "无生产组" in grouped_styles and grouped_styles["无生产组"]:
                st.write("### 无生产组的款式")
                no_group_data = []
                for style in grouped_styles["无生产组"]:
                    no_group_data.append({
                        "生产顺序": style.get("production_order", "-"),
                        "款号": style["style_number"],
                        "工序": style["process_type"],
                        "缝纫开始日期": f"{style['sewing_start_date']} ({style.get('start_time_period', '上午')})",
                        "周期": style["cycle"],
                        "订单数量": style["order_quantity"],
                        "日产量": style["daily_production"]
                    })
                st.table(no_group_data)
                
        col1, col2, col3 = st.columns(3)
        
        with col1:
            if st.button("生成所有生产流程图"):
                # 根据用户选择决定是否重新排序
                if enable_sequential_production:
                    # 重新安排同一生产组内款式的缝纫开始时间
                    styles_to_process = rearrange_styles_by_production_group(st.session_state["all_styles"])
                else:
                    styles_to_process = st.session_state["all_styles"]
                    
                # 创建一个临时目录来存储图片
                with tempfile.TemporaryDirectory() as temp_dir:
                    # 生成所有图表
                    for style in styles_to_process:
                        sewing_start_time = datetime.combine(style["sewing_start_date"], datetime.min.time())
                        start_time_period = style.get("start_time_period", "上午")  # 获取上午/下午信息
                        schedule = calculate_schedule(
                            sewing_start_time, 
                            style["process_type"], 
                            style["cycle"], 
                            style["order_quantity"], 
                            style["daily_production"],
                            start_time_period
                        )
                        # 设置当前款号和生产组用于标题显示
                        st.session_state["style_number"] = style["style_number"]
                        st.session_state["production_group"] = style.get("production_group", "")
                        fig = plot_timeline(schedule, style["process_type"], style["cycle"])
                        
                        # 保存图片 - 简化文件名
                        production_group = style.get("production_group", "")
                        # Include production group in filename if available
                        if production_group:
                            filename = f"{style['style_number']}_{production_group}_{style['process_type']}.png"
                        else:
                            filename = f"{style['style_number']}_{style['process_type']}.png"
                        filepath = os.path.join(temp_dir, filename)
                        fig.savefig(filepath, dpi=300, bbox_inches='tight')
                        plt.close(fig)
                    
                    # 创建ZIP文件
                    zip_path = os.path.join(temp_dir, "生产流程时间表.zip")
                    with zipfile.ZipFile(zip_path, 'w') as zipf:
                        for file in os.listdir(temp_dir):
                            if file.endswith('.png'):
                                zipf.write(os.path.join(temp_dir, file), file)
                    
                    # 提供ZIP文件下载
                    with open(zip_path, "rb") as f:
                        st.download_button(
                            label="下载所有图片(ZIP)",
                            data=f,
                            file_name="生产流程时间表.zip",
                            mime="application/zip"
                        )
        
        with col2:
            if st.button("生成部门时间线图"):
                # 根据用户选择决定是否重新排序
                if enable_sequential_production:
                    # 重新安排同一生产组内款式的缝纫开始时间
                    styles_to_process = rearrange_styles_by_production_group(st.session_state["all_styles"])
                else:
                    styles_to_process = st.session_state["all_styles"]
                # 生成部门时间线图
                #zip_path = generate_department_wise_plots(st.session_state["all_styles"])
                zip_path = generate_department_wise_plots(styles_to_process)
                # 提供ZIP文件下载
                with open(zip_path, "rb") as f:
                    st.download_button(
                        label="下载部门时间线图(ZIP)",
                        data=f,
                        file_name="部门时间线图.zip",
                        mime="application/zip"
                    )
        # with col3:
        #     if st.button("生成Excel报表"):
        #         # 根据用户选择决定是否重新排序
        #         if enable_sequential_production:
        #             # 重新安排同一生产组内款式的缝纫开始时间
        #             styles_to_process = rearrange_styles_by_production_group(st.session_state["all_styles"])
        #         else:
        #             styles_to_process = st.session_state["all_styles"]
                
        #         # 生成Excel报表
        #         excel_path = generate_excel_report(styles_to_process)
                
        #         # 提供Excel文件下载
        #         with open(excel_path, "rb") as f:
        #             st.download_button(
        #                 label="下载Excel报表",
        #                 data=f,
        #                 file_name="生产计划报表.xlsx",
        #                 mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        #             )
        with col3:
            if st.button("生成Excel报表"):
                # 根据用户选择决定是否重新排序
                if enable_sequential_production:
                    # 重新安排同一生产组内款式的缝纫开始时间
                    styles_to_process = rearrange_styles_by_production_group(st.session_state["all_styles"])
                else:
                    styles_to_process = st.session_state["all_styles"]
                
                # 生成Excel报表
                excel_path = generate_excel_report(styles_to_process)
                
                # 提供Excel文件下载
                with open(excel_path, "rb") as f:
                    st.download_button(
                        label="下载Excel报表",
                        data=f,
                        file_name="生产计划报表.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )

    # 调整生产流程部分保持不变
    if "schedule" in st.session_state:
        st.subheader("调整生产流程")
        
        # 选择部门和步骤
        selected_dept = st.selectbox("选择部门:", list(st.session_state["schedule"].keys()))
        if selected_dept:
            delayed_step = st.selectbox("选择延误的工序:", list(st.session_state["schedule"][selected_dept].keys()))
            new_end_date = st.date_input("选择新的完成时间:", min_value=datetime.today().date())
            # 转换date为datetime
            new_end_time = datetime.combine(new_end_date, datetime.min.time())
            
            if st.button("调整生产时间"):
                st.session_state["schedule"] = adjust_schedule(
                    st.session_state["schedule"],
                    selected_dept,
                    delayed_step,
                    new_end_time
                )
                fig = plot_timeline(st.session_state["schedule"], selected_process, cycle)
                
                # Display the plot in Streamlit
                st.pyplot(fig)
                
                # Add download button for high-resolution image
                buf = io.BytesIO()
                fig.savefig(buf, format='png', dpi=300, bbox_inches='tight')
                buf.seek(0)
                st.download_button(
                    label="下载高分辨率图片",
                    data=buf,
                    file_name=f"{style_number}_{selected_process}.png",
                    mime="image/png"
                )
